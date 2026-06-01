import csv
import os
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_csv_to_xlsx(csv_path, xlsx_path):
    print(f"Reading CSV from {csv_path}...")
    if not os.path.exists(csv_path):
        print(f"Error: {csv_path} does not exist.")
        return None, None, None

    # We read the data first to use in both xlsx and html
    rows = []
    created_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    
    with open(csv_path, mode='r', encoding='utf-8-sig') as f:
        first_line = f.readline()
        if "Created At" in first_line:
            # Parse actual created time from first line if exists
            try:
                created_at = first_line.split("Created At: ")[1].strip()
            except Exception:
                pass
            f.readline() # Skip Completed At
            f.readline() # Skip File Path
            reader = csv.reader(f)
        else:
            f.seek(0)
            reader = csv.reader(f)
            
        for row in reader:
            if row:
                rows.append(row)

    if not rows:
        print("No data found in CSV.")
        return None, None, None

    headers = rows[0]
    data = rows[1:]

    # --- EXCEL GENERATION ---
    print(f"Generating Excel workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Command Inventory"
    ws.views.sheetView[0].showGridLines = True

    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate Blue
    data_font = Font(name=font_family, size=10, color="000000")
    code_font = Font(name="Consolas", size=9, color="334155")
    
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    odd_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    status_styles = {
        "active": {
            "fill": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="166534")
        },
        "unclear": {
            "fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="92400E")
        },
        "disabled": {
            "fill": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            "font": Font(name=font_family, size=10, bold=True, color="991B1B")
        },
        "legacy": {
            "fill": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
            "font": Font(name=font_family, size=10, italic=True, color="4B5563")
        }
    }

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_header
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28

    # Write Data
    for row_idx, row_data in enumerate(data, 2):
        row_fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
        ws.row_dimensions[row_idx].height = 24
        
        extended_row_data = row_data + [""] * (len(headers) - len(row_data))
        for col_idx, value in enumerate(extended_row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = data_font
            cell.fill = row_fill
            
            col_name = headers[col_idx - 1].lower()
            if col_name in ["category", "command", "type"]:
                cell.alignment = align_center
                if col_name == "command":
                    cell.font = Font(name=font_family, size=10, bold=True, color="1E3A8A")
            elif col_name in ["entrypoint", "implementation"]:
                cell.alignment = align_left
                cell.font = code_font
            elif col_name == "status":
                cell.alignment = align_center
                val_lower = str(value).lower()
                matched = False
                for k, style in status_styles.items():
                    if k in val_lower:
                        cell.fill = style["fill"]
                        cell.font = style["font"]
                        matched = True
                        break
                if not matched:
                    cell.fill = row_fill
                    cell.font = data_font
            else:
                cell.alignment = align_left

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        col_name = headers[col[0].column - 1].lower()
        
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val_str) + 4)
            else:
                max_len = max(max_len, len(val_str))
        
        if col_name == "description":
            ws.column_dimensions[col_letter].width = 45
        elif col_name == "notes":
            ws.column_dimensions[col_letter].width = 40
        elif col_name in ["entrypoint", "implementation"]:
            ws.column_dimensions[col_letter].width = 35
        elif col_name in ["options_or_args"]:
            ws.column_dimensions[col_letter].width = 30
        elif col_name in ["permissions"]:
            ws.column_dimensions[col_letter].width = 25
        else:
            calculated_width = max(12, min(max_len + 3, 40))
            ws.column_dimensions[col_letter].width = calculated_width

    wb.save(xlsx_path)
    print(f"Created styled XLSX at: {xlsx_path}")

    # Return data for HTML generation
    return headers, data, created_at

def generate_html_dashboard(headers, data, created_at, html_path):
    print(f"Generating HTML dashboard...")
    # Convert data rows to list of dicts for JSON
    records = []
    categories = set()
    types = set()
    statuses = set()

    for row in data:
        record = {}
        for idx, col_name in enumerate(headers):
            val = row[idx] if idx < len(row) else ""
            record[col_name.lower()] = val
            
            if col_name.lower() == "category" and val:
                categories.add(val)
            elif col_name.lower() == "type" and val:
                types.add(val)
            elif col_name.lower() == "status" and val:
                statuses.add(val)
                
        records.append(record)

    # Sort categories, types, statuses
    categories_sorted = sorted(list(categories))
    types_sorted = sorted(list(types))
    statuses_sorted = sorted(list(statuses))

    # Stats calculation
    total_commands = len(records)
    slash_count = sum(1 for r in records if r.get('type') == 'slash')
    prefix_count = sum(1 for r in records if r.get('type') == 'prefix')
    alias_count = sum(1 for r in records if r.get('type') == 'alias')
    active_count = sum(1 for r in records if 'active' in r.get('status', '').lower())
    unclear_count = sum(1 for r in records if 'unclear' in r.get('status', '').lower())

    # Build the HTML content
    html_content = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Discord Bot Commands Inventory</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-color: #0b0f19;
            --card-bg: #151d30;
            --accent: #6366f1;
            --accent-glow: rgba(99, 102, 241, 0.15);
            --border-color: rgba(255, 255, 255, 0.06);
            --text-main: #f3f4f6;
            --text-muted: #9ca3af;
            
            --active-bg: rgba(16, 185, 129, 0.12);
            --active-text: #34d399;
            --unclear-bg: rgba(245, 158, 11, 0.12);
            --unclear-text: #fbbf24;
            --disabled-bg: rgba(239, 68, 68, 0.12);
            --disabled-text: #f87171;
            --legacy-bg: rgba(107, 114, 128, 0.12);
            --legacy-text: #9ca3af;
        }}

        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        body {{
            background-color: var(--bg-color);
            color: var(--text-main);
            font-family: 'Outfit', sans-serif;
            min-height: 100vh;
            padding: 1.5rem;
            line-height: 1.5;
            -webkit-font-smoothing: antialiased;
        }}

        header {{
            max-width: 1400px;
            margin: 0 auto 1.5rem auto;
            display: flex;
            flex-direction: column;
            gap: 0.5rem;
            position: relative;
        }}

        .header-top {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            flex-wrap: wrap;
            gap: 1rem;
        }}

        h1 {{
            font-size: 2.2rem;
            font-weight: 800;
            background: linear-gradient(135deg, #a5b4fc 0%, #6366f1 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: -0.025em;
        }}

        .metadata {{
            font-size: 0.875rem;
            color: var(--text-muted);
            background: rgba(255, 255, 255, 0.03);
            padding: 0.4rem 0.8rem;
            border-radius: 9999px;
            border: 1px solid var(--border-color);
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
        }}

        /* Dashboard Overview Grid */
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
            max-width: 1400px;
            margin: 0 auto 2rem auto;
        }}

        .stat-card {{
            background: var(--card-bg);
            border-radius: 16px;
            padding: 1.25rem;
            border: 1px solid var(--border-color);
            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.25);
            display: flex;
            flex-direction: column;
            gap: 0.25rem;
            transition: transform 0.2s ease, border-color 0.2s ease;
        }}

        .stat-card:hover {{
            transform: translateY(-2px);
            border-color: rgba(99, 102, 241, 0.3);
        }}

        .stat-card .label {{
            font-size: 0.875rem;
            font-weight: 500;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        .stat-card .value {{
            font-size: 2.25rem;
            font-weight: 800;
            color: #ffffff;
            display: flex;
            align-items: baseline;
            gap: 0.5rem;
        }}

        .stat-card.active-card {{ border-left: 4px solid #10b981; }}
        .stat-card.unclear-card {{ border-left: 4px solid #f59e0b; }}
        .stat-card.slash-card {{ border-left: 4px solid #818cf8; }}
        .stat-card.prefix-card {{ border-left: 4px solid #f472b6; }}

        /* Main Workspace Container */
        .workspace {{
            max-width: 1400px;
            margin: 0 auto;
            background: var(--card-bg);
            border-radius: 20px;
            border: 1px solid var(--border-color);
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.3);
            overflow: hidden;
            display: flex;
            flex-direction: column;
        }}

        /* Filters Bar */
        .filter-bar {{
            padding: 1.5rem;
            background: rgba(255, 255, 255, 0.01);
            border-bottom: 1px solid var(--border-color);
            display: flex;
            flex-wrap: wrap;
            gap: 1rem;
            align-items: center;
        }}

        .search-wrapper {{
            flex: 1;
            min-width: 280px;
            position: relative;
        }}

        .search-input {{
            width: 100%;
            padding: 0.8rem 1rem 0.8rem 2.8rem;
            background: rgba(11, 15, 25, 0.6);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            color: var(--text-main);
            font-family: inherit;
            font-size: 1rem;
            transition: all 0.2s ease;
        }}

        .search-input:focus {{
            outline: none;
            border-color: var(--accent);
            box-shadow: 0 0 0 3px var(--accent-glow);
        }}

        .search-icon {{
            position: absolute;
            left: 1rem;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-muted);
            pointer-events: none;
            width: 18px;
            height: 18px;
        }}

        .filter-select {{
            padding: 0.8rem 2rem 0.8rem 1rem;
            background: rgba(11, 15, 25, 0.6);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            color: var(--text-main);
            font-family: inherit;
            font-size: 0.95rem;
            cursor: pointer;
            outline: none;
            appearance: none;
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' fill='none' viewBox='0 0 24 24' stroke='%239ca3af'%3E%3Cpath stroke-linecap='round' stroke-linejoin='round' stroke-width='2' d='M19 9l-7 7-7-7'/%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: right 0.75rem center;
            background-size: 1rem;
            min-width: 180px;
            transition: border-color 0.2s ease;
        }}

        .filter-select:focus {{
            border-color: var(--accent);
        }}

        .tabs {{
            display: flex;
            background: rgba(11, 15, 25, 0.6);
            padding: 0.25rem;
            border-radius: 12px;
            border: 1px solid var(--border-color);
        }}

        .tab-btn {{
            padding: 0.55rem 1.25rem;
            border: none;
            background: transparent;
            color: var(--text-muted);
            font-family: inherit;
            font-size: 0.9rem;
            font-weight: 600;
            border-radius: 8px;
            cursor: pointer;
            transition: all 0.2s ease;
        }}

        .tab-btn.active {{
            background: var(--accent);
            color: #ffffff;
        }}

        /* Table Design for Desktop */
        .table-container {{
            overflow-x: auto;
            display: block;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            font-size: 0.95rem;
        }}

        th {{
            background: rgba(11, 15, 25, 0.4);
            color: #ffffff;
            font-weight: 600;
            padding: 1rem 1.25rem;
            border-bottom: 2px solid var(--border-color);
            text-transform: uppercase;
            font-size: 0.75rem;
            letter-spacing: 0.05em;
            white-space: nowrap;
        }}

        td {{
            padding: 1rem 1.25rem;
            border-bottom: 1px solid var(--border-color);
            vertical-align: middle;
        }}

        tr:last-child td {{
            border-bottom: none;
        }}

        tr {{
            transition: background-color 0.15s ease;
        }}

        tr:hover {{
            background-color: rgba(255, 255, 255, 0.02);
        }}

        /* Typography & Custom Cells */
        .cmd-name {{
            font-family: 'JetBrains Mono', monospace;
            font-weight: 700;
            color: #818cf8;
            font-size: 1rem;
            background: rgba(129, 140, 248, 0.08);
            padding: 0.2rem 0.5rem;
            border-radius: 6px;
            border: 1px solid rgba(129, 140, 248, 0.15);
            display: inline-block;
        }}

        .type-badge {{
            font-size: 0.75rem;
            font-weight: 700;
            text-transform: uppercase;
            padding: 0.25rem 0.6rem;
            border-radius: 9999px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
        }}

        .type-badge.slash {{ background: rgba(99, 102, 241, 0.15); color: #818cf8; border: 1px solid rgba(99, 102, 241, 0.3); }}
        .type-badge.prefix {{ background: rgba(236, 72, 153, 0.15); color: #f472b6; border: 1px solid rgba(236, 72, 153, 0.3); }}
        .type-badge.alias {{ background: rgba(156, 163, 175, 0.15); color: #d1d5db; border: 1px solid rgba(156, 163, 175, 0.3); }}

        .status-badge {{
            font-size: 0.75rem;
            font-weight: 700;
            text-transform: uppercase;
            padding: 0.25rem 0.6rem;
            border-radius: 6px;
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
        }}

        .status-badge.active {{ background: var(--active-bg); color: var(--active-text); }}
        .status-badge.unclear {{ background: var(--unclear-bg); color: var(--unclear-text); }}
        .status-badge.disabled {{ background: var(--disabled-bg); color: var(--disabled-text); }}
        .status-badge.legacy {{ background: var(--legacy-bg); color: var(--legacy-text); }}

        .code-cell {{
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.85rem;
            color: var(--text-muted);
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            max-width: 250px;
        }}

        .code-cell:hover {{
            color: #ffffff;
        }}

        /* Mobile Layout cards */
        .mobile-list {{
            display: none;
            flex-direction: column;
            gap: 1rem;
            padding: 1.25rem;
        }}

        .mobile-card {{
            background: rgba(255, 255, 255, 0.02);
            border-radius: 14px;
            border: 1px solid var(--border-color);
            padding: 1.25rem;
            display: flex;
            flex-direction: column;
            gap: 0.75rem;
            transition: all 0.2s ease;
        }}

        .mobile-card:active {{
            transform: scale(0.98);
            background: rgba(255, 255, 255, 0.04);
        }}

        .mobile-card-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 0.5rem;
        }}

        .mobile-row {{
            display: flex;
            flex-direction: column;
            gap: 0.25rem;
        }}

        .mobile-label {{
            font-size: 0.75rem;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        .mobile-value {{
            font-size: 0.95rem;
        }}

        .mobile-code {{
            font-family: 'JetBrains Mono', monospace;
            font-size: 0.8rem;
            background: rgba(0, 0, 0, 0.2);
            padding: 0.4rem 0.6rem;
            border-radius: 8px;
            border: 1px solid var(--border-color);
            word-break: break-all;
            color: #e2e8f0;
        }}

        .empty-state {{
            padding: 4rem 2rem;
            text-align: center;
            color: var(--text-muted);
            font-size: 1.1rem;
            display: none;
        }}

        /* Mobile responsiveness media queries */
        @media (max-width: 1024px) {{
            .table-container {{
                display: none;
            }}
            .mobile-list {{
                display: flex;
            }}
            .filter-bar {{
                flex-direction: column;
                align-items: stretch;
            }}
            .search-wrapper, .filter-select, .tabs {{
                width: 100%;
            }}
            .tabs {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                text-align: center;
            }}
            .tab-btn {{
                padding: 0.6rem 0;
            }}
            h1 {{
                font-size: 1.8rem;
            }}
        }}

        @media (max-width: 480px) {{
            body {{
                padding: 0.75rem;
            }}
            .stats-grid {{
                grid-template-columns: repeat(2, 1fr);
                gap: 0.75rem;
            }}
            .stat-card {{
                padding: 0.8rem;
            }}
            .stat-card .value {{
                font-size: 1.6rem;
            }}
            .stat-card .label {{
                font-size: 0.7rem;
            }}
            h1 {{
                font-size: 1.5rem;
            }}
            .metadata {{
                font-size: 0.75rem;
            }}
        }}
    </style>
</head>
<body>

    <header>
        <div class="header-top">
            <h1>Command Inventory</h1>
            <div class="metadata">
                <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"></circle><polyline points="12 6 12 12 16 14"></polyline></svg>
                <span>Cập nhật: {created_at}</span>
            </div>
        </div>
    </header>

    <div class="stats-grid">
        <div class="stat-card">
            <div class="label">Tổng số lệnh</div>
            <div class="value" id="stat-total">{total_commands}</div>
        </div>
        <div class="stat-card active-card">
            <div class="label">Đang hoạt động</div>
            <div class="value" id="stat-active">{active_count}</div>
        </div>
        <div class="stat-card unclear-card">
            <div class="label">Chưa rõ trạng thái</div>
            <div class="value" id="stat-unclear">{unclear_count}</div>
        </div>
        <div class="stat-card slash-card">
            <div class="label">Slash commands</div>
            <div class="value" id="stat-slash">{slash_count}</div>
        </div>
    </div>

    <div class="workspace">
        <div class="filter-bar">
            <div class="search-wrapper">
                <svg class="search-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"></circle><line x1="21" y1="21" x2="16.65" y2="16.65"></line></svg>
                <input type="text" id="search-input" class="search-input" placeholder="Tìm kiếm lệnh, mô tả, file code...">
            </div>

            <select id="category-select" class="filter-select">
                <option value="all">Tất cả danh mục ({len(categories_sorted)})</option>
                {"".join(f'<option value="{c}">{c}</option>' for c in categories_sorted)}
            </select>

            <div class="tabs">
                <button class="tab-btn active" data-type="all">Tất cả</button>
                <button class="tab-btn" data-type="slash">Slash</button>
                <button class="tab-btn" data-type="prefix">Prefix</button>
                <button class="tab-btn" data-type="alias">Alias</button>
            </div>
        </div>

        <div class="table-container">
            <table id="commands-table">
                <thead>
                    <tr>
                        <th style="width: 15%">Danh mục</th>
                        <th style="width: 15%">Lệnh</th>
                        <th style="width: 10%">Loại</th>
                        <th style="width: 25%">Mô tả</th>
                        <th style="width: 10%">Trạng thái</th>
                        <th style="width: 25%">File Implementation</th>
                    </tr>
                </thead>
                <tbody id="table-body">
                    <!-- Dynamic rendering -->
                </tbody>
            </table>
        </div>

        <div class="mobile-list" id="mobile-list">
            <!-- Dynamic rendering -->
        </div>

        <div class="empty-state" id="empty-state">
            Không tìm thấy lệnh nào khớp với bộ lọc của bạn.
        </div>
    </div>

    <script>
        // Data injected from CSV parser
        const data = {json.dumps(records, ensure_ascii=False)};

        // DOM elements
        const searchInput = document.getElementById('search-input');
        const categorySelect = document.getElementById('category-select');
        const tabButtons = document.querySelectorAll('.tab-btn');
        const tableBody = document.getElementById('table-body');
        const mobileList = document.getElementById('mobile-list');
        const emptyState = document.getElementById('empty-state');

        // Stats elements
        const statTotal = document.getElementById('stat-total');
        const statActive = document.getElementById('stat-active');
        const statUnclear = document.getElementById('stat-unclear');
        const statSlash = document.getElementById('stat-slash');

        let currentFilters = {{
            search: '',
            category: 'all',
            type: 'all'
        }};

        // Set up tab switching listeners
        tabButtons.forEach(btn => {{
            btn.addEventListener('click', () => {{
                tabButtons.forEach(b => b.classList.remove('active'));
                btn.classList.add('active');
                currentFilters.type = btn.dataset.type;
                filterAndRender();
            }});
        }});

        // Input listeners
        searchInput.addEventListener('input', (e) => {{
            currentFilters.search = e.target.value.toLowerCase().trim();
            filterAndRender();
        }});

        categorySelect.addEventListener('change', (e) => {{
            currentFilters.category = e.target.value;
            filterAndRender();
        }});

        function filterAndRender() {{
            const filtered = data.filter(item => {{
                // Type filter
                if (currentFilters.type !== 'all' && item.type !== currentFilters.type) {{
                    return false;
                }}
                
                // Category filter
                if (currentFilters.category !== 'all' && item.category !== currentFilters.category) {{
                    return false;
                }}
                
                // Search filter
                if (currentFilters.search) {{
                    const searchStr = [
                        item.command,
                        item.description,
                        item.implementation,
                        item.entrypoint,
                        item.category,
                        item.permissions,
                        item.notes,
                        item.options_or_args
                    ].join(' ').toLowerCase();
                    
                    if (!searchStr.includes(currentFilters.search)) {{
                        return false;
                    }}
                }}
                
                return true;
            }});

            renderTable(filtered);
            renderMobile(filtered);
            updateStats(filtered);

            if (filtered.length === 0) {{
                emptyState.style.display = 'block';
            }} else {{
                emptyState.style.display = 'none';
            }}
        }}

        function updateStats(filteredList) {{
            statTotal.innerText = filteredList.length;
            
            const active = filteredList.filter(item => item.status && item.status.toLowerCase().includes('active')).length;
            const unclear = filteredList.filter(item => item.status && item.status.toLowerCase().includes('unclear')).length;
            const slash = filteredList.filter(item => item.type === 'slash').length;

            statActive.innerText = active;
            statUnclear.innerText = unclear;
            statSlash.innerText = slash;
        }}

        function getStatusClass(status) {{
            const s = (status || '').toLowerCase();
            if (s.includes('active')) return 'active';
            if (s.includes('unclear')) return 'unclear';
            if (s.includes('disabled')) return 'disabled';
            if (s.includes('legacy')) return 'legacy';
            return 'unclear';
        }}

        function renderTable(items) {{
            tableBody.innerHTML = '';
            
            items.forEach(item => {{
                const statusClass = getStatusClass(item.status);
                const typeClass = (item.type || 'slash').toLowerCase();
                const displayStatus = item.status || 'unknown';
                
                const tr = document.createElement('tr');
                tr.innerHTML = `
                    <td><span style="color: var(--text-muted); font-size: 0.85rem;">${{item.category || ''}}</span></td>
                    <td><span class="cmd-name">${{item.command || ''}}</span></td>
                    <td><span class="type-badge ${{typeClass}}">${{item.type || 'slash'}}</span></td>
                    <td style="font-size: 0.9rem;">
                        <div style="font-weight: 500;">${{item.description || ''}}</div>
                        ${{item.options_or_args ? `<div style="font-size: 0.75rem; color: var(--text-muted); margin-top: 0.25rem;">Args: <code>${{item.options_or_args}}</code></div>` : ''}}
                        ${{item.notes ? `<div style="font-size: 0.75rem; color: #f59e0b; font-style: italic; margin-top: 0.25rem;">Chú ý: ${{item.notes}}</div>` : ''}}
                    </td>
                    <td><span class="status-badge ${{statusClass}}">${{displayStatus}}</span></td>
                    <td>
                        <div class="code-cell" title="${{item.implementation || ''}}">${{item.implementation || ''}}</div>
                        ${{item.entrypoint ? `<div style="font-size: 0.75rem; color: var(--text-muted); font-family: monospace;" title="${{item.entrypoint}}">Trigger: ${{item.entrypoint}}</div>` : ''}}
                    </td>
                `;
                tableBody.appendChild(tr);
            }});
        }}

        function renderMobile(items) {{
            mobileList.innerHTML = '';
            
            items.forEach(item => {{
                const statusClass = getStatusClass(item.status);
                const typeClass = (item.type || 'slash').toLowerCase();
                const displayStatus = item.status || 'unknown';
                
                const card = document.createElement('div');
                card.className = 'mobile-card';
                
                let detailsHtml = '';
                if (item.description) {{
                    detailsHtml += `
                        <div class="mobile-row">
                            <span class="mobile-label">Mô tả</span>
                            <span class="mobile-value" style="font-weight: 500;">${{item.description}}</span>
                        </div>
                    `;
                }}
                
                if (item.options_or_args) {{
                    detailsHtml += `
                        <div class="mobile-row">
                            <span class="mobile-label">Options/Args</span>
                            <span class="mobile-code">${{item.options_or_args}}</span>
                        </div>
                    `;
                }}

                if (item.implementation) {{
                    detailsHtml += `
                        <div class="mobile-row">
                            <span class="mobile-label">Implementation</span>
                            <span class="mobile-code">${{item.implementation}}</span>
                        </div>
                    `;
                }}
                
                if (item.notes) {{
                    detailsHtml += `
                        <div class="mobile-row">
                            <span class="mobile-label">Ghi chú</span>
                            <span class="mobile-value" style="color: #fbbf24; font-style: italic; font-size: 0.85rem;">${{item.notes}}</span>
                        </div>
                    `;
                }}
                
                card.innerHTML = `
                    <div class="mobile-card-header">
                        <span class="cmd-name">${{item.command || ''}}</span>
                        <div style="display: flex; gap: 0.4rem;">
                            <span class="type-badge ${{typeClass}}">${{item.type || 'slash'}}</span>
                            <span class="status-badge ${{statusClass}}">${{displayStatus}}</span>
                        </div>
                    </div>
                    <div class="mobile-row" style="margin-bottom: 0.25rem;">
                        <span class="mobile-label">Danh mục</span>
                        <span class="mobile-value" style="color: var(--text-muted); font-size: 0.85rem;">${{item.category || 'Chưa phân loại'}}</span>
                    </div>
                    ${{detailsHtml}}
                `;
                mobileList.appendChild(card);
            }});
        }}

        // Initial render
        filterAndRender();
    </script>
</body>
</html>
"""
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"Created styled HTML dashboard at: {html_path}")

if __name__ == "__main__":
    csv_file = r"C:\Bot Discord\reports\bot-command-inventory.csv"
    xlsx_file = r"C:\Bot Discord\reports\bot-command-inventory.xlsx"
    html_file = r"C:\Bot Discord\reports\bot-command-inventory.html"
    
    headers, data, created_at = convert_csv_to_xlsx(csv_file, xlsx_file)
    if headers and data:
        generate_html_dashboard(headers, data, created_at, html_file)
